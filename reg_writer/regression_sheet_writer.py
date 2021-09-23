import openpyxl


class RegressionSheetWriter:
    def __init__(self, path):
        self.path = path
        self.workbook = openpyxl.load_workbook(path)
        self.active_sheet = self.workbook.active

        self.UID_COL = 1
        self.RESULT_COL = 5
        self.NOTES_COL = 6

        self.reg_uid_dict = {}

    def load_regression_file(self):
        """ Load the regression file and populate the uid dictionary """
        max_row = self.active_sheet.max_row

        for i in range(1, max_row + 1):
            cell = self.active_sheet.cell(row=i, column=self.UID_COL)
            if cell.value and cell.value.startswith('REG_'):
                self.reg_uid_dict[cell.value] = i


    def write_test_result(self, reg_uid, result, message):
        # Get the row for the reg_uid
        # Probably need some error handling around here if this returns None (reg_uid doesnt exist)
        reg_row = self.reg_uid_dict.get(reg_uid)

        pass_or_fail = 'Pass' if result else 'Fail'

        # Write result to sheet
        self.active_sheet.cell(column=self.RESULT_COL, row=reg_row, value=pass_or_fail)

        # Write message to sheet
        self.active_sheet.cell(column=self.NOTES_COL, row=reg_row, value=message)

        # Save sheet
        self.workbook.save(self.path)